import os
sheet_path = os.environ.get('sheet_path')

from bottle import route,run,request, template, static_file
from openpyxl import load_workbook
global sheet
def query_sheet(id):
    wb = load_workbook(filename =sheet_path)
    sheet = wb['all']
    vals = []
    match_col = 2
    match_val = id.lower()
    last_run = False
    print match_val
    titles = [ cell[0].value for cell in sheet.iter_cols() ]
    max_cols = len(titles)
    print max_cols
    range_start = 0
    range_end = max_cols
    for row in sheet.iter_rows():
        print row[match_col].value
        if (row[match_col].value is not None and str(row[match_col].value).lower() == match_val):
            #print "entered"
            vals += ([{"title":titles[i],"value":row[i].value} for i in range(range_start,range_end) ])
            #if last_run:
                #break
            # match_col = 7
            # match_val = "eramanaged"
            # range_start = 7
            # range_end = 14
            # last_run = True
            # vals += ([{"title":"","value":""},{"title":"","value":""}])

    if vals:
        vals[0]["title"] = "Attendee Name"
        f = lambda a: a or ''
        vals[0]["value"] = f(vals[0]["value"]) +" " + f(vals[1]["value"])
        del vals[1]
        del vals[1]
        print vals
    return vals



@route('/home')
def home_page():
  print request.url
  return template("home.html",srcurl=(request.url).split('/')[0])

@route('/contact')
def contact_page():
  return template("contact.html",srcurl=(request.url).split('/')[0])

@route('/')
def entry():
    return template("form1.html",rows=[],error="",srcurl=(request.url).split('/')[0])

@route('/',method="POST")
def formhandler():
    attn_id = request.forms.get("attn_id")
    print attn_id
    headers={"accept":"application/json"}
    itemdict=query_sheet(attn_id.strip())
    print itemdict
    if not itemdict:
        return template("form1.html",rows=itemdict,error="No entries found",srcurl=(request.url).split('/')[0])
    else:
        return template("form1.html",rows=itemdict,error="",srcurl=(request.url).split('/')[0])


run(server='cherrypy',host='0.0.0.0',port=8090,debug=True)
